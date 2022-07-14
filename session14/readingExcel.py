import openpyxl

# File → Workbook → Sheets → Rows → Cells
file = "C:\SeleniumPractice\data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows = sheet.max_row    # count number of rows
cols = sheet.max_column # count number of columns

# read all the rows & columns
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end='\t\t')
    print()